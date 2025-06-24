import json
from openpyxl import Workbook, load_workbook
import os

def read_sheets_file(file_name):
    if not os.path.exists(file_name):
        wb = Workbook()

        ws1 = wb.active
        ws1.title = "websites"
        ws1.append(["WEBSITE", "LINKS", "EMAILS"])
        # ws1.append(["www.wooga.com", "", ""])
        # ws1.append(["www.crux-consulting.ai", "", ""])

        ws2 = wb.create_sheet(title="emails")
        ws2.append(["EMAIL", "DONE", "NAME", "SURNAME", "GENDER", "TITLE", "POSITION", "COMPANY", "ADDRESS", "TELEPHONE", "", "", "", "", "", "LINKEDIN", "LINKS"])

        wb.save(file_name)
    else:
        wb = load_workbook(file_name)
    websites_dict = load_sheet_data(wb["websites"])
    emails_dict = load_sheet_data(wb["emails"])
    return wb, websites_dict, emails_dict


def load_sheet_data(sheet):
    out = {}
    for row in sheet.iter_rows(min_row=2, values_only=False):
        name = row[0].value
        if name:
            out[name] = row
    return out

def update_website_data(wb, website_dict, website, links, emails):
    if website in website_dict:
        row = website_dict[website]
        row[1].value = json.dumps(links)
        row[2].value = json.dumps(emails)
    else:
        new_row = wb['websites'].append([website, json.dumps(links), json.dumps(emails)])
        website_dict[website] = new_row

links_column_index = 16  # Column Q is the 17th column (0-indexed)
def update_email_links_data(wb, emails_dict, email, links):
    if email in emails_dict:
        row = emails_dict[email]
        existing_links = json.loads(row[links_column_index].value) if row[links_column_index].value else []
        combined_links = list(set(existing_links + links))
        row[links_column_index].value = json.dumps(combined_links)
    else:
        new_row = wb['emails'].append([email] + [None]*(links_column_index-1) + [json.dumps(links)])
        emails_dict[email] = new_row

data_keys_to_rows = {
    "first_name": 2,
    "last_name": 3,
    "gender": 4,
    "title": 5,
    "position": 6,
    # "company": 7,
}
def update_email_person_data(wb, emails_dict, email, data):

    if email in emails_dict:
        row = emails_dict[email]
        count_done = 0
        for key, index in data_keys_to_rows.items():
            if row[index].value is None or len(row[index].value) == 0:
                if key in data and data[key] is not None:
                    row[index].value = data[key]
                    count_done += 1
            else:
                count_done += 1
        if count_done == len(data_keys_to_rows):
            row[1].value = "Y"
def get_email_links(emails_dict, email):
    if email in emails_dict:
        row = emails_dict[email]
        links = row[links_column_index].value
        linkedin_link = row[links_column_index-1].value
        return (json.loads(links) if links else []) + ([linkedin_link] if linkedin_link else [])
    return []